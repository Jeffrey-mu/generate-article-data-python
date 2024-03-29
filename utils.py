import openpyxl
import requests
import pypandoc
import json
import re


def stringEncodingFun(string):
    if not string:
        return string
    specialCharacters = r'&nbsp;|&emsp;|&ensp;|(<!--.*?-->)'
    string = re.sub(r'&#40;', '(', string)
    string = re.sub(r'&#41;', ')', string)
    string = re.sub(r'&#34;', '\"', string)
    string = re.sub(r'&#39;', '\'', string)
    string = re.sub(r'&#8194;', '', string)
    string = re.sub(r'&#8195;', '', string)
    string = re.sub(r'&#160;', '', string)
    string = re.sub(r'&#60;', '<', string)
    string = re.sub(r'&#62;', '>', string)
    string = re.sub(r'&#38;', '&', string)
    string = re.sub(r'&#34;', '\"', string)
    string = re.sub(r'&#169;', '©', string)
    string = re.sub(r'&#174;', '®', string)
    string = re.sub(r'&#8482;', '™', string)
    string = re.sub(r'&#215;', '×', string)
    string = re.sub(r'&#247;', '÷', string)
    string = re.sub(r'&nbsp;', ' ', string)
    string = re.sub(r'&emsp;', '', string)
    string = re.sub(r'&ensp;', '', string)
    string = re.sub(r'&lt;', '<', string)
    string = re.sub(r'&gt;', '>', string)
    string = re.sub(r'&amp;', '&', string)
    string = re.sub(r'&quot;', '\"', string)
    string = re.sub(r'&copy;', '©', string)
    string = re.sub(r'&reg;', '®', string)
    string = re.sub(r'&times;', '×', string)
    string = re.sub(r'&divide;', '÷', string)
    string = re.sub(r'(<!--.*?-->)', '', string)
    string = re.sub(r'\\t', '', string)
    string = re.sub(r'\'', '"', string)
    return string


def format_html(data, file_name):
    html = data
    pypandoc.convert_text(html, 'docx', 'html', outputfile=file_name)  # 将 html 代码转化成docx


# 处理文本
def revise_html(html_str, overall=True):
    if html_str == "" or not isinstance(html_str, str):
        return ""
    # 定义匹配的正则表达式
    pattern = re.compile(r"^<(p|h1|h2|h3|li)>.*<\/\1>$")
    pattern_img = re.compile(r"^<img\s.*\s?\/?>$")
    html_str = html_str.replace('<div>', '')
    html_str = html_str.replace('</div>', '')
    # 剔除所有div
    # 定义一个空列表，用于存放处理后的每个元素
    wrapped_arr = []
    # 循环遍历每个元素并添加包裹符号
    for elem in list(filter(lambda x: x.strip() != "", html_str.split('\n'))):
        if not html_str.__contains__('<p>') and not html_str.__contains__('ul>') and not elem.__contains__('overall'):
            wrapped_arr.append(f"<p>{elem}</p>")
        else:
            result = pattern.match(elem) or pattern_img.match(elem) or elem.__contains__('ul>')
            wrapped_elem = elem
            if result:
                wrapped_arr.append(wrapped_elem)
        # 使用join()方法将所有元素合并成一个字符串
        embroidery_data = "\n".join(wrapped_arr)
    return embroidery_data


def read_excel(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 获取工作表名称
    sheet_name = workbook.sheetnames[0]

    # 获取工作表对象
    sheet = workbook[sheet_name]

    # 定义一个空列表
    data = []

    # 获取标题行的值，用于作为字典的键
    headers = [cell.value for cell in sheet[1]]

    # 遍历每一行，并将每一行的数据整理成字典形式，添加到列表中
    for row in sheet.iter_rows(min_row=2):
        # 定义一个空字典，用于存储当前行的数据
        row_data = {}
        for index, cell in enumerate(row):
            # 将单元格的值添加到当前行的数据字典中
            row_data[headers[index]] = cell.value
        # 将当前行的数据字典添加到整个数据列表中
        data.append(row_data)

    # 打印整个数据列表
    return data


def get_atticl_data(title):
    response = requests.get(
        f"http://47.104.212.164:3000/dataList?data_type_id=&main_title={title}&content=&src=&pageIndex=1&pageSize=10&manager_name=admin_plus",
        headers={
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh,zh-CN;q=0.9,en;q=0.8",
            "authorization": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJuYW1lIjoiYWRtaW5fcGx1cyIsInBhc3N3b3JkIjoiaWRndGVjaG5ldHdvcmstODA4IiwiaWF0IjoxNjc4OTMyNjMyLCJleHAiOjE2NzkwMTkwMzJ9.iCa1-hYSjCbNaEUjFRr9VNWLHS2aaVxK8lK1Z1M4kVI"
        })
    stream = response.json()
    return stream


def is_json(json_str):
    try:
        json.loads(json_str)
    except ValueError as e:
        return False
    return True


def revise_json(json_str):
    pattern = re.compile(r'"content": \[(.*?)\]', re.DOTALL)
    match = pattern.search(json_str)
    if match:
        content_str = match.group(1)
        arr = content_str.split('\n')
        for i in range(len(arr) - 1):
            if arr[i] != '' and not '{' in arr[i] and not '}' in arr[i] and not '}' in arr[i + 1]:
                if not arr[i].endswith(','):
                    arr[i] = arr[i] + ','
        json_str = json_str.replace(content_str, '\n'.join(arr))
    else:
        print("No match found.")
    return json_str


def cut_json(json_str):
    pattern = re.compile(r'\{(.+)\}', re.DOTALL)
    match = pattern.search(json_str)
    return '{' + match.group(1) + '}'
